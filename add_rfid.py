#!/usr/bin/env python
# -*- coding: utf-8 -*-

from app import db
from app.model import Rfid_ids
import openpyxl
db.create_all()

filepath = 'ww.xlsx'
wb = openpyxl.load_workbook(filepath, data_only=True)
ws = wb.active


def fio_to_fi(item):
    fio = item.split()
    if not fio:
        return None
    return f'{fio[0].capitalize()} {fio[1][0]}.'


for i in range(2, ws.max_row+1):
    rfid_id = ws.cell(row=i, column=2).value
    fio = ws.cell(row=i, column=1).value
    fi = fio_to_fi(fio)
    db.session.add(Rfid_ids(rfid_id, fi))
    print(rfid_id, fio, '_', fi)


db.session.commit()
