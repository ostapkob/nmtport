import pandas as pd
from sqlalchemy import create_engine
try:
    from middleware import  list_mechanisms 
except ModuleNotFoundError:
    import  list_mechanisms 
from datetime import datetime, timedelta
from rich import print
from openpyxl import Workbook
import smtplib
import csv
from tabulate import tabulate
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import pickle
import time

import os
import sys
import inspect
current_dir = os.path.dirname(os.path.abspath(
    inspect.getfile(inspect.currentframe())))
parent_dir = os.path.dirname(current_dir)
sys.path.insert(0, parent_dir)
from psw import mail_pass


HOST = 'smtp.yandex.ru'
FROM ='smartportdaly@yandex.ru'

nameTerminal = {1: "УТ-1", 2: "ГУТ-2"}
addresses = {
    1: [
        'Vadim.Evsyukov@nmtport.ru'
        'Maxim.Anufriev@nmtport.ru'
        'Konstantin.Nikitenko@nmtport.ru'

        'Pavel.Shunin@nmtport.ru',
        'Oleg.Evsyukov@nmtport.ru', 
        'Disp.Smen@nmtport.ru',
        # 'Vladimir.Grigoriev@nmtport.ru',
        # 'Dmitry.Chernyavskiy@nmtport.ru',
        # 'Radion.Bespalov@nmtport.ru',
        'Petr.Gerasimenko@nmtport.ru',
        'Alexander.Ostapchenko@nmtport.ru',
        'ostap666@yandex.ru'

        ],
    2: [
        'Dmitry.Golynsky@nmtport.ru'
        'Vyacheslav.Gaz@nmtport.ru'
        'Vladimir.Speransky@nmtport.ru'
        'Denis.Medvedev@nmtport.ru'

        'Pavel.Shunin@nmtport.ru',
        'Oleg.Evsyukov@nmtport.ru', 
        'Disp.Smen@nmtport.ru',
        # 'Vladimir.Grigoriev@nmtport.ru',
        'Petr.Gerasimenko@nmtport.ru',
        'Alexander.Ostapchenko@nmtport.ru',
        'ostap666@yandex.ru'
    ],
}
titles = {
    1: [
    "номер крана ", 
    "начало смены </br>(> 8:20)", 
    "окончание перед обедом </br>(< 12:00)", 
    "начало после обеда </br>(> 13:00)", 
    "окончание перед тех. перерывом </br>(< 16:30)", 
    'начало после тех. перерыва </br>(> 17:00)', 
    'окончание смены </br>(< 19:40)',
    'общие потери по крану </br> (минут)',
    ],
    2: [
    "номер крана ", 
    "начало смены </br>(> 20:20)", 
    "окончание перед обедом </br>(< 01:00)", 
    "начало после обеда </br>(> 02:00)", 
    "окончание перед тех. перерывом </br>(< 04:30)", 
    'начало после тех. перерыва </br>(> 05:00)', 
    'окончание смены </br>(< 07:40)',
    'общие потери по крану </br> (минут)',
    ]
}

krans = list_mechanisms.kran
usm = list_mechanisms.usm

krans_UT = [45, 34, 53, 69, 21, 37, 4, 41, 5, 36, 40, 32, 25, 11, 33, 20, 8, 22, 12, 13, 6, 26, 47, 54, 14, 16, 82]
krans_GUT = [28, 18, 1, 35, 31, 17, 58, 60, 49, 38, 39, 23, 48, 72, 65, 10]

def getData(mech_id, date_shift, shift ):
    mech_id = str(mech_id)
    date_shift= str(date_shift) 
    shift = str(shift)
    ServerName = "192.168.99.106"
    Database = "nmtport"
    UserPwd = "ubuntu:Port2020"
    Driver = "driver=ODBC Driver 17 for SQL Server"
    engine = create_engine('mssql+pyodbc://' + UserPwd + '@' + ServerName + '/' + Database + "?" + Driver)
    sql ="""
    SELECT TOP (1000) [id]
          ,[mechanism_id]
          ,[value]
          ,dateadd(hour, 10, [timestamp]) as time
          ,[date_shift]
          ,[shift]
          ,[terminal]
    FROM [nmtport].[dbo].[post]
    where 
    mechanism_id=""" + mech_id + """ and 
    date_shift='"""  + date_shift + """' and
    shift=""" + shift + """ 
    order by timestamp  """

    df = pd.read_sql(sql, engine)
    df = df.set_index('time')
    return df


def get_yellow_diapozones (date, shift):
    tommorow = date + timedelta(days=1)
    date = str(date) + " "
    tommorow = str(tommorow) + " "
    if shift == 1:
        diapozones = {
        "start" :        [date + '08:00', date + '08:20'],
        "work_1" :       [date + '08:20', date + '11:58'],
        "lanch_start" :  [date + '11:58', date + '12:00'],
        "lanch_finish" : [date + '13:00', date + '13:02'],
        "work_2" :       [date + '13:02', date + '16:28'],
        "tea_start" :    [date + '16:28', date + '16:30'],
        "tea_finish" :   [date + '17:00', date + '17:02'],
        "work_3" :       [date + '17:02', date + '19:40'],
        "finish" :       [date + '19:40', date + '20:00'],
        }
    if shift == 2:
        diapozones = {
        "start" :        [date + '20:00',     date + '20:20'],
        "work_1" :       [date + '20:20',     tommorow + '00:58'],
        "lanch_start" :  [tommorow + '00:58', tommorow + '01:00'],
        "lanch_finish" : [tommorow + '02:00', tommorow + '02:02'],
        "work_2" :       [tommorow + '02:02', tommorow + '04:28'],
        "tea_start" :    [tommorow + '04:28', tommorow + '04:30'],
        "tea_finish" :   [tommorow + '05:00', tommorow + '05:02'],
        "work_3" :       [tommorow + '05:02', tommorow + '07:40'],
        "finish" :       [tommorow + '07:40', tommorow + '08:00'],
        }
    return diapozones

def sum_diapozone(diapozone):
    summ = 0
    for x in diapozone['value']:
        if x>1:
            summ += 1
    return summ

def diapozone(df, diapozone):
    return df[diapozone[0] : diapozone[1]]


def get_mech_diapozones(df, diapozones):
    return {
        "start" : diapozone(df, diapozones['start']),
        "work_1" : diapozone(df, diapozones['work_1']),
        "lanch_start" : diapozone(df, diapozones['lanch_start']),
        "lanch_finish" : diapozone(df, diapozones['lanch_finish']),
        "work_2" : diapozone(df, diapozones['work_2']),
        "tea_start" : diapozone(df, diapozones['tea_start']),
        "tea_finish" : diapozone(df, diapozones['tea_finish']),
        "work_3" : diapozone(df, diapozones['work_3']),
        "finish" : diapozone(df, diapozones['finish']),
    }

def sum_mech_diapozones(diapozones):
    return { num : sum_diapozone(diapozone) for  num, diapozone in diapozones.items()}

def get_time(mech_sum, mech_diapozones, condition):
    work_zone = 20
    work, period, position = condition
    if mech_sum[work] > work_zone and not mech_sum[period]:
        side = mech_diapozones[work].query('value > 0').iloc[position]
        return side.name 
    return None

def get_hour_and_minutes(time):
    if time:
        h=time.hour
        m=time.minute
        if h<10:
            h = '0' + str(h)
        if m<10:
            m = '0' + str(m)
        return f'{h}:{m}'
    return '' 

def convert_time_to_str(times):
    return [get_hour_and_minutes(time) for time in times]

def save_to_xlsx(list_kran, name):
    wb=Workbook()
    wb.create_sheet('kran_work_time')
    del wb['Sheet']
    ws = wb['kran_work_time']
    titles = ["number", "start", "start_lanc", "finish_lanch", "start_tea", 'finish_tea', 'finish']
    for en, title in enumerate(titles, 1):
        ws.cell(row=1, column=en, value=title)
    for en, item in enumerate(list_kran, 2):
        for col, value in enumerate(item, 1):
            ws.cell(row=en, column=col, value=value)
    wb.save(filename = 'kran_periods'+name+'.xlsx')

def find_periods(date, shift, terminal):
    diapozones = get_yellow_diapozones(date, shift)
    mechanisms = {} 
    search_conditions = (
        ('work_1', 'start', 0),
        ('work_1', 'lanch_start', -1),
        ('work_2', 'lanch_finish', 0),
        ('work_2', 'tea_start', -1),
        ('work_3', 'tea_finish', 0),
        ('work_3', 'finish', -1)
    )
    if terminal==1:
        krans_terminal = [(k, v) for k, v in krans.items() if k in krans_UT]
    else:
        krans_terminal = [(k, v) for k, v in krans.items() if k in krans_GUT]

    for kran_num, kran_id in krans_terminal:
        df = getData(kran_id, date, shift)
        mech_diapozones = get_mech_diapozones(df, diapozones)
        mech_sum = sum_mech_diapozones(mech_diapozones)
        mech_zones = []

        for condition in search_conditions:
            mech_zones.append(get_time(mech_sum, mech_diapozones, condition))

        if any(mech_zones):
            mechanisms[kran_num] = mech_zones

    return mechanisms

def get_bg(time, red_zones, en):

    if time is None:
        return 'bg-white'
    if en in [0, 2, 4]:
        if time > red_zones[en]:
            return 'bg-red'
    if en in [1, 3, 5]:
        if time < red_zones[en]:
            return 'bg-red'
    return 'bg-yellow'

def get_red_zones(date, shift):
    format = '%Y-%m-%d %H:%M'
    diapozones = get_yellow_diapozones(date, shift)
    condition = [
        datetime.strptime(diapozones['start'][1], format) + timedelta(minutes=10), # >
        datetime.strptime(diapozones['lanch_start'][0], format) - timedelta(minutes=5), # <
        datetime.strptime(diapozones['lanch_finish'][1], format) + timedelta(minutes=5), # >
        datetime.strptime(diapozones['tea_start'][0], format) - timedelta(minutes=5), # <
        datetime.strptime(diapozones['tea_finish'][1], format) + timedelta(minutes=5), # >
        datetime.strptime(diapozones['finish'][0], format) - timedelta(minutes=10), # <
    ]
    return condition

def get_border_zones(date, shift):
    format = '%Y-%m-%d %H:%M'
    diapozones = get_yellow_diapozones(date, shift)
    condition = [
        datetime.strptime(diapozones['start'][1], format) ,
        datetime.strptime(diapozones['lanch_start'][1], format), 
        datetime.strptime(diapozones['lanch_finish'][0], format),
        datetime.strptime(diapozones['tea_start'][1], format),
        datetime.strptime(diapozones['tea_finish'][0], format),
        datetime.strptime(diapozones['finish'][0], format),
    ]
    return condition


def add_bg(times, red_zones):
    time_and_bg = []
    for en, time in enumerate(times):
        time_and_bg.append((get_bg(time, red_zones, en), time))
    return time_and_bg

def make_table(data, date, shift):
    if not data:
        return None
    red_zones = get_red_zones(date, shift)
    border_zones = get_border_zones(date, shift)
    total = 0
    table = '<h4>' +  str(shift) + " смена </h4>  <table>" 
    for cell in titles[shift]:
        table += f'<td class=titles> {cell} </td>'
    table += '</tr> <tr>'
    for kran_num, values in data.items():
        table += ' <td class=number>  ' + str(kran_num) + ' </td>'
        row = add_bg(values, red_zones)
        count_sum = count_different(values, border_zones)
        total += count_sum
        for bg, time in row:
            value = get_hour_and_minutes(time)
            table += f'<td class={bg}> {value} </td>'
        table += f'<td class=sum>' + str(count_sum) + '</td>'
        table += '</tr>'
    table += '<tr>' + '<td class=empty></td>'*7 + '<td class=total>' + str(total) +  '</td></tr>'
    table += '</table>' 
    return table

def count_different(real_time, border_zones):
    assert len(border_zones) == len(real_time)
    sum_time = 0
    for i in range(len(border_zones)):
        if real_time[i] is not None:
            dt = (real_time[i] - border_zones[i]).total_seconds()
            sum_time += abs(dt/60)
    return int(sum_time)

def make_html(table1, table2, date):
    table1 = table1 if table1  else ""
    table2 = table2 if table2  else ""

    html = """
    <html>
        <head>
        <style> 
          table, th, td {{ border: 1px solid #999; border-collapse: collapse; }}
          th, td {{ padding: 5px; }}
          .bg-red {{
            background: #F4A9A9;
          }}
          .bg-yellow {{
            background: #FFFFA4;
          }}
          .bg-white {{
            background: #FFFFFF;
          }}

          .titles {{
            background: #F5F5F5;
            color: #444;
          }}
          .number {{
            background: #F9F9F9;
            text-align: center;
            font-weight: bold;
          }}
          .sum {{
            font-weight: bold;
            text-align: right;
            color: #666;
          }}
          .empty {{
            border: 0px solid #fff;
          }}
          .total {{
            font-weight: bold;
            text-align: right;
            color: #111;
          }}
        </style>
        </head>
        <body>
            <p>""" + str(date) + """</p>
            <p>Позднее начало, ранее окончание по 
            производственным периодам</p>""" + table1 +  """ 
            </br>""" + table2 +  """ 
            </br>
            <a href="https://m1.nmtport.ru/krans"> SmartPort </a>
        </body>
    </html>
    """
    return html
def mail_to_str_list(mail):
    msg = ""
    for  m in mail:
        msg += m + '; '
    return msg

def sent_email(periods1, periods2, date, terminal):
    tmp1 =  [[k, *convert_time_to_str(v)] for k,v in periods1.items()]
    tmp2 =  [[k, *convert_time_to_str(v)] for k,v in periods2.items()]
    data = tmp1 +[['-']*7]+ tmp1
    SUBJECT = f"простои кранов {nameTerminal[terminal]} {str(date)}"
    text = str(date) + """
    Позднее начало, ранее окончание по производственным периодам
    {table}
    SmartPort
    """
    table1 = make_table(periods1, date, 1)
    table2 = make_table(periods2, date, 2)
    html = make_html(table1, table2, date)

    text = text.format(table=tabulate(data, headers="firstrow", tablefmt="grid"))
    html = html.format(table=tabulate(data, headers="firstrow", tablefmt="html"))
    TO = ''    

    message = MIMEMultipart(
    "alternative", None, [MIMEText(text), MIMEText(html,'html')])
    message['Subject'] = SUBJECT
    message['From'] = FROM
    message['To'] = 'Alexander.Ostapchenko@nmtport.ru'

    # with open('mail.html', 'w') as f:
    #     f.write(html)
    #     f.close()
    server = smtplib.SMTP_SSL(HOST, 465)
    server.ehlo()
    server.login(FROM, mail_pass)
    server.sendmail(FROM, addresses[terminal], message.as_string())
    server.quit()

def every_day():
    yesterday =  datetime.now().date() - timedelta(days=1)

    UT_shift_1 = find_periods(yesterday, 1, 1)
    UT_shift_2 = find_periods(yesterday, 2, 1)
    GUT_shift_1 = find_periods(yesterday, 1, 2)
    GUT_shift_2 = find_periods(yesterday, 2, 2)
    sent_email(UT_shift_1, UT_shift_2, yesterday, 1)
    sent_email(GUT_shift_1, GUT_shift_2, yesterday, 2)

if __name__ == "__main__":

    # yesterday =  datetime.now().date() - timedelta(days=1)
    # UT_shift_1 = find_periods(yesterday, 1, 1)
    # UT_shift_2 = find_periods(yesterday, 2, 1)
    # GUT_1_shift_1 = find_periods(yesterday, 1, 2)
    # GUT_2_shift_2 = find_periods(yesterday, 2, 2)
    
    # name_file_pickle = 'dump'+'_'+str(yesterday)+"_"+str(1)
    # with open(name_file_pickle, 'wb') as f:
    #     pickle.dump(UT_shift_1, f)
    # name_file_pickle = 'dump'+'_'+str(yesterday)+"_"+str(2)
    # with open(name_file_pickle, 'wb') as f:
    #     pickle.dump(UT_shift_2, f)

    # name_file_pickle = 'dump'+'_'+str(yesterday)+"_"+str(1)
    # with open(name_file_pickle, 'rb') as f:
    #     UT_shift_1 = pickle.load(f)
    # name_file_pickle = 'dump'+'_'+str(yesterday)+"_"+str(3)
    # with open(name_file_pickle, 'rb') as f:
    #     UT_shift_2 = pickle.load(f)

    # sent_email(UT_shift_1, UT_shift_2, yesterday, 1)


    while True:
        hour = datetime.now().hour
        minute = datetime.now().minute
        if hour==10 and minute==0:
            print(datetime.now())
            every_day()
            time.sleep(60)
        time.sleep(15)



