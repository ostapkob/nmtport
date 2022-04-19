import pandas as pd
from sqlalchemy import create_engine
try:
    from middleware import  list_mechanisms 
except ModuleNotFoundError:
    import  list_mechanisms 
from datetime import datetime, timedelta
from rich import print
from openpyxl import Workbook
import smtplib
import csv
from tabulate import tabulate
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

# HOST = 'mail.nmtport.ru'
HOST = 'smtp.yandex.ru'
TO = 'Alexander.Ostapchenko@nmtport.ru'
FROM = 'ostap666@yandex.ru'
# FROM ='office-rocket@yandex.ru'
# FROM ='smartportdaly@yandex.ru'

passw = ''
nameTerminal = {1: 'УТ-1 ', 2: "ГУТ-2"}
addresses = {
    1: [
        'Vadim.Evsyukov@nmtport.ru'
        'Konstantin.Nikitenko@nmtport.ru'
        'Maxim.Anufriev@nmtport.ru'
        'Dmitry.Chernyavskiy@nmtport.ru'
        ],
    2: [
        'Dmitry.Golynsky@nmtport.ru'
        'Vyacheslav.Gaz@nmtport.ru'
        'Vladimir.Speransky@nmtport.ru'
        'Denis.Medvedev@nmtport.ru'
    ],
    3: [
        'Alexander.Ostapchenko@nmtport.ru',
        'Petr.Gerasimenko@nmtport.ru',
        'Radion.Bespalov@nmtport.ru'
    ],
}
# passw = 'Port=2022'
# passw = 'or11235813'

krans = list_mechanisms.kran
usm = list_mechanisms.usm

krans_UT = [45, 34, 53, 69, 21, 37, 4, 41, 5, 36, 40, 32, 25, 11, 33, 20, 8, 22, 12, 13, 6, 26, 47, 54, 14, 16, 82]
krans_GUT = [28, 18, 1, 35, 31, 17, 58, 60, 49, 38, 39, 23, 48, 72, 65, 10]

def getData(mech_id, date_shift, shift ):
    mech_id = str(mech_id)
    date_shift= str(date_shift) 
    shift = str(shift)
    ServerName = "192.168.99.106"
    Database = "nmtport"
    UserPwd = "ubuntu:Port2020"
    Driver = "driver=ODBC Driver 17 for SQL Server"
    engine = create_engine('mssql+pyodbc://' + UserPwd + '@' + ServerName + '/' + Database + "?" + Driver)
    sql ="""
    SELECT TOP (1000) [id]
          ,[mechanism_id]
          ,[value]
          ,dateadd(hour, 10, [timestamp]) as time
          ,[date_shift]
          ,[shift]
          ,[terminal]
    FROM [nmtport].[dbo].[post]
    where 
    mechanism_id=""" + mech_id + """ and 
    date_shift='"""  + date_shift + """' and
    shift=""" + shift + """ 
    order by timestamp  """

    df = pd.read_sql(sql, engine)
    df = df.set_index('time')
    return df


def get_red_diapozones (date, shift):
    tommorow = date + timedelta(days=1)
    date = str(date) + " "
    tommorow = str(tommorow) + " "
    if shift == 1:
        diapozones = {
        "start" :        [date + '08:00', date + '08:20'],
        "work_1" :       [date + '08:20', date + '11:50'],
        "lanch_start" :  [date + '11:50', date + '12:00'],
        "lanch_finish" : [date + '13:00', date + '13:10'],
        "work_2" :       [date + '13:10', date + '15:50'],
        "tea_start" :    [date + '15:50', date + '16:00'],
        "tea_finish" :   [date + '17:00', date + '17:10'],
        "work_3" :       [date + '17:10', date + '19:15'],
        "finish" :       [date + '19:15', date + '20:00'],
        }
    if shift == 2:
        diapozones = {
        "start" :        [date + '20:00',     date + '20:20'],
        "work_1" :       [date + '20:20',     tommorow + '00:50'],
        "lanch_start" :  [tommorow + '00:50', tommorow + '01:00'],
        "lanch_finish" : [tommorow + '02:00', tommorow + '02:10'],
        "work_2" :       [tommorow + '02:10', tommorow + '03:50'],
        "tea_start" :    [tommorow + '03:50', tommorow + '04:00'],
        "tea_finish" :   [tommorow + '05:00', tommorow + '05:10'],
        "work_3" :       [tommorow + '05:10', tommorow + '07:15'],
        "finish" :       [tommorow + '07:15', tommorow + '08:00'],
        }
    return diapozones

def sum_diapozone(diapozone):
    summ = 0
    for x in diapozone['value']:
        if x>1:
            summ += 1
    return summ

def diapozone(df, diapozone):
    return df[diapozone[0] : diapozone[1]]


def get_mech_diapozones(df, diapozones):
    return {
        "start" : diapozone(df, diapozones['start']),
        "work_1" : diapozone(df, diapozones['work_1']),
        "lanch_start" : diapozone(df, diapozones['lanch_start']),
        "lanch_finish" : diapozone(df, diapozones['lanch_finish']),
        "work_2" : diapozone(df, diapozones['work_2']),
        "tea_start" : diapozone(df, diapozones['tea_start']),
        "tea_finish" : diapozone(df, diapozones['tea_finish']),
        "work_3" : diapozone(df, diapozones['work_3']),
        "finish" : diapozone(df, diapozones['finish']),
    }

def sum_mech_diapozones(diapozones):
    return { num : sum_diapozone(diapozone) for  num, diapozone in diapozones.items()}

def get_time(mech_sum, mech_diapozones, condition):
    work_zone = 20
    work, period, position = condition
    if mech_sum[work] > work_zone and not mech_sum[period]:
        first = mech_diapozones[work].query('value > 0').iloc[position]
        return first.name 
    return None

def get_hour_and_minutes(time):
    if time:
        h=time.hour
        m=time.minute
        if h<9:
            h = '0' + str(h)
        if m<9:
            m = '0' + str(m)
        return f'{h}:{m}'
    return None

def save_to_xlsx(list_kran, name):
    wb=Workbook()
    wb.create_sheet('kran_work_time')
    del wb['Sheet']
    ws = wb['kran_work_time']
    titles = ["number", "start", "start_lanc", "finish_lanch", "start_tea", 'finish_tea', 'finish']
    for en, title in enumerate(titles, 1):
        ws.cell(row=1, column=en, value=title)
    for en, item in enumerate(list_kran, 2):
        for col, value in enumerate(item, 1):
            ws.cell(row=en, column=col, value=value)
    wb.save(filename = 'kran_periods'+name+'.xlsx')

def find_periods(date, shift, terminal):
    diapozones = get_red_diapozones(date, shift)
    list_mechanisms = []
    search_conditions = (
        ('work_1', 'start', 0),
        ('work_1', 'lanch_start', -1),
        ('work_2', 'lanch_finish', 0),
        ('work_2', 'tea_start', -1),
        ('work_3', 'tea_finish', 0),
        ('work_3', 'finish', -1)
    )
    if terminal==1:
        krans_terminal = [(k, v) for k, v in krans.items() if k in krans_UT]
    else:
        krans_terminal = [(k, v) for k, v in krans.items() if k in krans_GUT]

    for kran_num, kran_id in krans_terminal:
        df = getData(kran_id, date, shift)
        mech_diapozones = get_mech_diapozones(df, diapozones)
        mech_sum = sum_mech_diapozones(mech_diapozones)
        mech_zones = []

        for condition in search_conditions:
            mech_zones.append(get_time(mech_sum, mech_diapozones, condition))
        mech_zones = [get_hour_and_minutes(time) for time in mech_zones]

        if any(mech_zones):
            list_mechanisms.append([kran_num, *mech_zones])

    return list_mechanisms


def make_table(data, shift):
    if not data:
        return None
    table = '<h5>' +  str(shift) + " смена </h5>  <table>" 
    titles = [
        "номер крана", 
        "начало смены", 
        "окончание перед обедом", 
        "начало после обеда ", 
        "окончание перед тех. перерывом", 
        'начало после тех. перерыва', 
        'окончание смены'
        ]
    data.insert(0,titles)
    for row in data:
        table += '<tr>'
        for cell in row:
            cell = '' if cell is None else cell
            table += f'<td>{cell}</td>'
        table += '</tr>'
    table += '</table>' 
    return table

def make_html(table1, table2,date):
    html = """
    <html>
    <head>
    <style> 
      table, th, td {{ border: 1px solid #999; border-collapse: collapse; }}
      th, td {{ padding: 5px; }}
    </style>
    </head>
    <body><p>""" + date + """</p>
    <p>Позднее начало, ранее окончание по 
    производственным периодам</p>""" + table1 +  """ 
    </br>""" + table2 +  """ 
    </br>
    <a href="https://m1.nmtport.ru/krans"> SmartPort </a>
    </body></html>
    """
    return html


def sent_email(periods1, periods2, date, terminal):
    date = str(date)
    data = periods1 +['-']*7+ periods2
    SUBJECT = "простои кранов "+ nameTerminal[terminal]  + date
    text = date + """
    Позднее начало, ранее окончание по производственным периодам
    {table}
    SmartPort
    """
    table1 = make_table(periods1, 1)
    table2 = make_table(periods2, 2)
    html = make_html(table1, table2, date)

    text = text.format(table=tabulate(data, headers="firstrow", tablefmt="grid"))
    html = html.format(table=tabulate(data, headers="firstrow", tablefmt="html"))

    message = MIMEMultipart(
    "alternative", None, [MIMEText(text), MIMEText(html,'html')])
    message['Subject'] = SUBJECT
    message['From'] = FROM
    message['To'] = TO

    server = smtplib.SMTP_SSL(HOST, 465)
    server.ehlo()
    server.login(FROM, passw)
    server.sendmail(FROM, [TO], message.as_string())
    server.quit()


if __name__ == "__main__":

    yesterday =  datetime.now().date() - timedelta(days=1)
    UT_shift_1 = find_periods(yesterday, 1, 1)
    UT_shift_2 = find_periods(yesterday, 2, 1)
    # GUT_1 = find_periods(yesterday, 1, 2)
    # GUT_2 = find_periods(yesterday, 2, 2)
    sent_email(UT_shift_1, UT_shift_2, yesterday, 1)
    # save_to_xlsx(UT_1, 'UT')
    # save_to_xlsx(GUT_1, 'GUT')



