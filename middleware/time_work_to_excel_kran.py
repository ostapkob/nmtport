from pymongo import MongoClient
from datetime import datetime, timedelta
from pprint import pp
from collections import defaultdict
from openpyxl import Workbook

client = MongoClient('mongodb://localhost:27017')
mongodb = client['HashShift']

krans_UT = [45, 34, 53, 69, 21, 37, 4, 41, 5, 36, 40, 32, 25, 11, 33, 20, 8, 22, 12, 13, 6, 26, 47, 54, 14, 16, 82]
krans_GUT = [28, 18, 1, 35, 31, 17, 58, 60, 49, 38, 39, 23, 48, 72, 65, 10]


def diapozone_time_work(date_start, date_finish, type_mechanism ):
    date_shift = date_start
    result = defaultdict(dict)
    while date_shift <= date_finish:
        for shift in range(1,3):
            date_str = date_shift.strftime("%d.%m.%Y")
            mongo_request = mongodb[type_mechanism].find_one({"_id": f"{date_str}|{shift}"})
            if not mongo_request:
                result[ date_shift.strftime("%Y-%m-%d") ] [shift] = {}
                continue
            del mongo_request["_id"]
            result[ date_shift.strftime("%Y-%m-%d") ] [shift] = mongo_request
        date_shift += timedelta(days=1)
    return result


def get_sum_time_work(mech_data):
    sum_not_work = 0
    for interval in mech_data.values():
        if interval['step']>15 and interval['value'] < 1: #mech stay or not power > 15 minutes
            sum_not_work += interval['step']
    if sum_not_work <5:
        return 0
    return  720-sum_not_work #720 = 60 * 12 hours


def avg(intervals):
    if intervals:
        return round(sum(intervals) / len(intervals))
    return None

def get_avg_by_terminals(data_by_days):
    '''if not values return None'''
    result = []
    for date in data_by_days.keys():
        for shift in range(1,3):
            mech_UT = []
            mech_GUT = []

            for mech in data_by_days[date][shift].values():
                mech_data = mech['data']
                num = mech['number']
                if num in krans_UT:
                    mech_UT.append(get_sum_time_work(mech_data))
                elif num in krans_GUT:
                    mech_GUT.append(get_sum_time_work(mech_data))
            result.append([date, shift, 'УТ-1',  avg(mech_UT) ])
            result.append([date, shift, 'ГУТ-2',  avg(mech_GUT) ])
    return result

def get_detal_work_time(data_by_days):
    '''if not values return None'''
    result = []
    for date in data_by_days.keys():
        for shift in range(1,3):
            mech_UT = []
            mech_GUT = []
            for mech in data_by_days[date][shift].values():
                mech_data = mech['data']
                num = mech['number']
                degrees90 = mech['total_90']
                degrees180 = mech['total_180']
                fio = mech['fio']
                contract = mech['contract']
                start = mech.get('start', None)
                finish = mech.get('finish', None)
                grab = mech.get('grab', None)

                if num in krans_UT:
                    ter = 'УТ-1'
                elif num in krans_GUT:
                    ter = 'ГУТ-2'
                result.append([
                    date, 
                    shift, 
                    ter,
                    num,
                    get_sum_time_work(mech_data),
                    degrees90,
                    degrees180,
                    fio,
                    contract,
                    start,
                    finish,
                    grab
                ])
    return result

def save_to_xlsx(avg_by_terminals):
    wb=Workbook()
    wb.create_sheet('kran_work_time')
    del wb['Sheet']
    ws = wb['kran_work_time']
    titles = ["date", "shift", "terminal", "number", "time", 'degrees90', 'degrees180', 'fio', 'contract', 'start', 'finish', 'grab']
    for en, title in enumerate(titles, 1):
        ws.cell(row=1, column=en, value=title)
    for en, item in enumerate(avg_by_terminals, 2):
        date, shift, terminal, number, time, degrees90, degrees180, fio, contract, start, finish, grab = item
        ws.cell(row=en, column=1, value=date)
        ws.cell(row=en, column=2, value=shift)
        ws.cell(row=en, column=3, value=terminal)
        ws.cell(row=en, column=4, value=number)
        ws.cell(row=en, column=5, value=time)
        ws.cell(row=en, column=6, value=degrees90)
        ws.cell(row=en, column=7, value=degrees180)
        ws.cell(row=en, column=8, value=fio)
        ws.cell(row=en, column=9, value=contract)
        ws.cell(row=en, column=10, value=start)
        ws.cell(row=en, column=11, value=finish)
        ws.cell(row=en, column=12, value=grab)
    wb.save(filename = 'kran_work_time.xlsx')

if __name__ == "__main__":
    date_start  = datetime(2021, 9, 1).date()
    date_finish = datetime(2022, 4, 1).date()
    assert date_start <= date_finish, "start < finish"
    type_mechanism = 'kran'
    data_by_days=diapozone_time_work(date_start, date_finish, type_mechanism)
    # avg_by_terminals = get_avg_by_terminals(data_by_days)
    detal_work_time = get_detal_work_time(data_by_days)
    save_to_xlsx(detal_work_time)


