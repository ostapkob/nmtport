from pymongo import MongoClient
from datetime import datetime, timedelta
from pprint import pp
from collections import defaultdict
from openpyxl import Workbook

client = MongoClient('mongodb://localhost:27017')
mongodb = client['HashShift']


def diapozone_time_work(date_start, date_finish, type_mechanism):
    date_shift = date_start
    result = defaultdict(dict)
    while date_shift <= date_finish:
        for shift in range(1, 3):
            date_str = date_shift.strftime("%d.%m.%Y")
            mongo_request = mongodb[type_mechanism].find_one(
                {"_id": f"{date_str}|{shift}"})
            if not mongo_request:
                result[date_shift.strftime("%Y-%m-%d")][shift] = {}
                continue
            del mongo_request["_id"]
            result[date_shift.strftime("%Y-%m-%d")][shift] = mongo_request
        date_shift += timedelta(days=1)
    return result


def get_sum_time_work(mech_data):
    sum_not_work = 0
    for interval in mech_data.values():
        # mech stay or not power > 15 minutes
        if interval['step'] > 15 and interval['value'] < 1:
            sum_not_work += interval['step']
    if sum_not_work < 10:
        return 0
    return 720-sum_not_work  # 720 = 60 * 12 hours


def get_detal_work_time(data_by_days):
    '''if not values return None'''
    result = []
    for date in data_by_days:
        for shift in range(1, 3):
            mech_UT = []
            mech_GUT = []
            for mech in data_by_days[date][shift].values():
                mech_data = mech['data']
                num = mech['number']
                fio = mech['fio']
                contract = mech['contract']
                start = mech.get('start', None)
                finish = mech.get('finish', None)
                time_coal = mech.get('time_coal', None)
                tons_in_hour = mech.get('tons_in_hour', None)
                ter = 'УТ-1'
                result.append([
                    date,
                    shift,
                    ter,
                    num,
                    get_sum_time_work(mech_data),
                    fio,
                    contract,
                    start,
                    finish,
                    time_coal * tons_in_hour
                ])
    return result


def save_to_xlsx(avg_by_terminals):
    wb = Workbook()
    wb.create_sheet('usm_work_time')
    del wb['Sheet']
    ws = wb['usm_work_time']
    titles = ["date", "shift", "terminal", "number", "time",
              'fio', 'contract', 'start', 'finish', 'total_tons']
    for en, title in enumerate(titles, 1):
        ws.cell(row=1, column=en, value=title)
    for en, item in enumerate(avg_by_terminals, 2):
        date, shift, terminal, number, time, fio, contract, start, finish, total_tons = item
        ws.cell(row=en, column=1, value=date)
        ws.cell(row=en, column=2, value=shift)
        ws.cell(row=en, column=3, value=terminal)
        ws.cell(row=en, column=4, value=number)
        ws.cell(row=en, column=5, value=time)
        ws.cell(row=en, column=6, value=fio)
        ws.cell(row=en, column=7, value=contract)
        ws.cell(row=en, column=8, value=start)
        ws.cell(row=en, column=9, value=finish)
        ws.cell(row=en, column=10, value=total_tons)
    wb.save(filename='usm_work_time.xlsx')


if __name__ == "__main__":
    # date_start  = datetime(2021, 9, 1).date()
    date_start = datetime(2022, 9, 1).date()
    date_finish = datetime(2022, 9, 30).date()
    assert date_start <= date_finish, "start < finish"
    type_mechanism = 'usm'
    data_by_days = diapozone_time_work(date_start, date_finish, type_mechanism)
    detal_work_time = get_detal_work_time(data_by_days)
    save_to_xlsx(detal_work_time)
    # pp(detal_work_time)
